from django.http import HttpResponse
from openpyxl import load_workbook
import json, os

def open_excel_file(path):
    try:
        # Load the workbook
        workbook = load_workbook(filename=path, rich_text=True, read_only=True)
        sheet = workbook.active
        return sheet

    except FileNotFoundError:
        print(f"The file {path} was not found.")
    except openpyxl.utils.exceptions.InvalidFileException:
        print("The file 'example.xlsx' is not a valid Excel file.")
    except Exception as e:
        print(f"An error occurred: {e}")

def transform_excel_file(sheet):
    #read header
    headers = []
    count_header= 0
    for cell in sheet[1]:
        val = cell.value if cell.value is not None else ''
        headers.append(str(val))
        count_header = count_header + 1
        if count_header == 10:
            break 

    # read data 
    data_list = []
    for row in sheet.iter_rows(min_row=2, max_row=5, max_col=10): #strating from second row as 1st row is header
        row_val= []
        for cell in row:
            print(cell.value)
            val = cell.value if cell.value is not None else ''
            row_val.append(str(val))
        data_list.append(row_val)
    
    # ceating dict 
    data_dict = {}
    data_dict['headers'] = headers
    data_dict['value'] = data_list
    return data_dict

def read_document(url):
    data_sheet = open_excel_file(url)
    data = transform_excel_file(data_sheet)
    print(data)
    print(type(data))
    return data


# exel_file_path = '.static/the-office-lines.xlsx'
# read_document(exel_file_path)